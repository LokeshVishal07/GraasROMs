"""Warehouse Bulk Upload: Excel template generation, validation, error
reporting, and applying validated rows to ReturnOrder + Evidence."""
import datetime as dt
import io

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from lib.db import get_session, ReturnOrder
from lib.evidence import save_link, is_probably_valid_link
from lib.utils import log_activity, push_notification

FONT_NAME = "Arial"

SYSTEM_COLUMNS = [
    "order_id", "return_id", "inspection_status", "warehouse_comments",
    "damage_reason", "image_drive_link", "inspection_date", "warehouse_user_name",
]
HUMAN_LABELS = {
    "order_id": "Order ID *", "return_id": "Return ID",
    # No longer starred/required -- see CASES_NOT_REQUIRING_INSPECTION below.
    "inspection_status": "Inspection Status", "warehouse_comments": "Warehouse Comments",
    "damage_reason": "Damage Reason", "image_drive_link": "Image / Drive Link",
    "inspection_date": "Inspection Date", "warehouse_user_name": "Warehouse User Name",
}
# inspection_status is no longer unconditionally required -- see
# _status_required_for() in validate_upload for the actual (order-aware) rule.
REQUIRED_COLUMNS = {"order_id"}

# Six outcomes now, not three: Refund Only, Request Cancelled, and Lost & Scrap
# cover cases where there's nothing to physically inspect.
STATUS_CHOICES_HUMAN = ["Good Condition", "Damaged", "Not Received", "Refund Only",
                         "Request Cancelled", "Lost & Scrap"]

# English + baseline Thai synonyms -> internal status code. This is a starting
# set (common warehouse phrasing), not an exhaustive list -- anything typed
# that ISN'T recognized here is no longer a hard rejection (see
# validate_upload): the row still gets applied, with the original text kept
# as the comment and flagged for someone to reclassify later, rather than
# blocking the whole upload over a status the dropdown doesn't happen to know.
STATUS_MAP = {
    "good condition": "good", "damaged": "damaged", "not received": "not_received",
    "refund only": "refund_only", "request cancelled": "request_cancelled",
    "request canceled": "request_cancelled", "lost & scrap": "lost_scrap",
    "lost and scrap": "lost_scrap", "lost": "lost_scrap", "scrap": "lost_scrap",
    # Thai synonyms seen in real uploads / commonly used by the warehouse team.
    "สภาพดี": "good", "เสียหาย": "damaged", "ของไม่ครบ": "damaged",
    "ไม่ได้รับสินค้า": "not_received", "ไม่ได้รับ": "not_received",
    "คืนเงินเท่านั้น": "refund_only", "ยกเลิกคำขอ": "request_cancelled",
    "สูญหาย": "lost_scrap", "ตัดจำหน่าย": "lost_scrap",
}
# These statuses are the ones where a photo/video normally matters -- but as
# of the warehouse team's feedback, a missing image no longer blocks the
# upload. Comments are still required (see the STATUS_REQUIRES_EVIDENCE check
# in validate_upload), but if the image/drive link is left blank the row still
# applies -- we just note "Image missing" in the order's comment so Operations
# can see it needs following up, rather than rejecting the whole row.
STATUS_REQUIRES_EVIDENCE = {"damaged", "not_received"}

# Deterministic code -> display label, for the upload preview table. Built
# from STATUS_CHOICES_HUMAN rather than reversing STATUS_MAP, since several
# STATUS_MAP keys (English + Thai synonyms) now point at the same code.
CODE_TO_LABEL = {STATUS_MAP[label.lower()]: label for label in STATUS_CHOICES_HUMAN}

# If the order this row resolves to already shows one of these trigger codes
# (set during the normal GRAAS import -- see classify_trigger in importer.py),
# the warehouse never actually received a parcel for it, so Inspection Status
# can be left blank without an error. When left blank for one of these, we
# auto-fill inspection_status to match rather than leaving it "pending".
CASES_NOT_REQUIRING_INSPECTION_BY_TRIGGER = {
    "cancelled_not_shipped": "request_cancelled",
    "cancelled_after_shipped": "request_cancelled",
}
# GRAAS's raw order_status for a parcel the 3PL lost in transit -- same idea,
# blank Inspection Status is fine and we infer Lost & Scrap.
CASES_NOT_REQUIRING_INSPECTION_BY_ORDER_STATUS = {
    "LOST_BY_3PL": "lost_scrap",
}

EXAMPLE_ROW = [
    "2201234567890", "RET-000123", "Damaged", "Box crushed, item cracked on arrival",
    "Physical damage in transit", "https://drive.google.com/file/d/example/view",
    "2026-07-19", "Wichai Warehouse",
]

MAX_TEMPLATE_ROWS = 300


# --------------------------------------------------------------------------
# Template
# --------------------------------------------------------------------------

def build_template_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Inspection Upload"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
    key_font = Font(name=FONT_NAME, italic=True, size=8, color="999999")
    example_font = Font(name=FONT_NAME, italic=True, color="808080", size=10)
    normal_font = Font(name=FONT_NAME, size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    yellow_fill = PatternFill("solid", fgColor="FFFF00")

    ws.append(SYSTEM_COLUMNS)
    for c in range(1, len(SYSTEM_COLUMNS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = key_font
        cell.alignment = Alignment(horizontal="center")

    ws.append([HUMAN_LABELS[k] for k in SYSTEM_COLUMNS])
    for c in range(1, len(SYSTEM_COLUMNS) + 1):
        cell = ws.cell(row=2, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.append(EXAMPLE_ROW)
    for c in range(1, len(SYSTEM_COLUMNS) + 1):
        cell = ws.cell(row=3, column=c)
        cell.font = example_font
        cell.border = border

    # order_id / return_id are long numeric-looking strings (TikTok order IDs
    # run 18-19 digits). Without this, Excel/Sheets "helpfully" reformats a
    # pasted value like 585061544671709159 into 5.8506E+17 -- and once that's
    # happened the original digits are gone for good, no amount of parsing on
    # our end can recover them. Pre-setting these two columns to Text before
    # anyone pastes into them is the actual fix; see also the apostrophe-
    # prefix fallback noted on the Instructions tab for manual entry.
    text_format_cols = {SYSTEM_COLUMNS.index("order_id") + 1, SYSTEM_COLUMNS.index("return_id") + 1}

    for r in range(4, 4 + MAX_TEMPLATE_ROWS):
        for c in range(1, len(SYSTEM_COLUMNS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = normal_font
            cell.border = border
            cell.fill = yellow_fill
            if c in text_format_cols:
                cell.number_format = "@"

    for c in text_format_cols:
        ws.cell(row=3, column=c).number_format = "@"

    widths = [18, 14, 16, 28, 22, 30, 16, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 12
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A4"

    status_col = SYSTEM_COLUMNS.index("inspection_status") + 1
    dv = DataValidation(type="list", formula1=f'"{",".join(STATUS_CHOICES_HUMAN)}"',
                         allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(status_col)}4:{get_column_letter(status_col)}{3 + MAX_TEMPLATE_ROWS}")

    # Instructions sheet
    ins = wb.create_sheet("Instructions")
    ins.column_dimensions["A"].width = 24
    ins.column_dimensions["B"].width = 72
    title_font = Font(name=FONT_NAME, size=14, bold=True, color="1F3864")
    bold = Font(name=FONT_NAME, size=11, bold=True)
    body = Font(name=FONT_NAME, size=11)

    ins["A1"] = "Warehouse Bulk Inspection Upload — Fill-in Guide"
    ins["A1"].font = title_font
    ins.merge_cells("A1:B1")

    ins["A3"] = "How this works"
    ins["A3"].font = bold
    ins["A4"] = ("Fill in one row per return on the 'Bulk Inspection Upload' tab — one order per row. "
                 "Delete the grey EXAMPLE row (row 3) before you finish. Fields marked with * are "
                 "mandatory. The file is validated on upload; any problem rows are rejected with a "
                 "downloadable error report, and only rows that pass validation are applied.")
    ins["A4"].alignment = Alignment(wrap_text=True)
    ins.merge_cells("A4:B4")
    ins.row_dimensions[4].height = 60

    ins["A5"] = ("⚠ Order ID and Return ID must stay as TEXT, not numbers. Columns A and B are "
                 "pre-formatted as Plain Text in this template, but pasting from another sheet can "
                 "still override that. If a long ID ever turns into something like 5.8506E+17, the "
                 "original digits are gone for good — before pasting, format the destination "
                 "columns as Plain Text (Format ▸ Number ▸ Plain text), or if typing an ID by hand, "
                 "prefix it with an apostrophe (e.g. '585061544671709159).")
    ins["A5"].font = Font(name=FONT_NAME, size=11, bold=True, color="B91C1C")
    ins["A5"].alignment = Alignment(wrap_text=True)
    ins.merge_cells("A5:B5")
    ins.row_dimensions[5].height = 75

    rows = [
        ("Column", "What to enter"),
        ("order_id", "REQUIRED. Must exactly match an Order ID already in the system. Keep as text — see the "
                      "warning above."),
        ("return_id", "The marketplace/DKSH return case number, if available. Helps disambiguate "
                       "when an Order ID appears more than once in the system, and when the same Order ID "
                       "legitimately repeats across rows (multiple returns, or one return split across SKUs)."),
        ("inspection_status", "One of: Good Condition, Damaged, Not Received, Refund Only, Request Cancelled, "
                               "Lost & Scrap (dropdown provided). Can be left BLANK if the order is already "
                               "known to be cancelled or lost — the system fills it in automatically from what "
                               "it already knows about that order. Thai equivalents are also accepted; if you "
                               "type something the system doesn't recognize, the row still goes through and "
                               "gets flagged for someone to double-check rather than being rejected."),
        ("warehouse_comments", "REQUIRED when status is Damaged or Not Received."),
        ("damage_reason", "Reason for the damage, if applicable."),
        ("image_drive_link", "A Google Drive / OneDrive / SharePoint link to photo or video evidence. "
                              "Recommended for Damaged or Not Received, but not required — you can upload "
                              "without it. If left blank, the system notes \"Image missing\" in that "
                              "return's comments so Operations knows to follow up for evidence."),
        ("inspection_date", "Date the item was inspected. Format: YYYY-MM-DD. Defaults to today if left blank."),
        ("warehouse_user_name", "Name of the warehouse staff who performed the inspection. Defaults to "
                                 "your logged-in name if left blank."),
    ]
    start_row = 7
    ins.cell(row=start_row, column=1, value=rows[0][0]).font = bold
    ins.cell(row=start_row, column=2, value=rows[0][1]).font = bold
    for i, (col, desc) in enumerate(rows[1:], start=1):
        r = start_row + i
        ins.cell(row=r, column=1, value=col).font = Font(name=FONT_NAME, size=11, bold=True)
        c2 = ins.cell(row=r, column=2, value=desc)
        c2.font = body
        c2.alignment = Alignment(wrap_text=True)
        ins.row_dimensions[r].height = 46

    note_row = start_row + len(rows) + 1
    ins.cell(row=note_row, column=1, value="Validation").font = bold
    ins.cell(row=note_row, column=2, value=(
        "Rows are rejected only if: Order ID is missing; the Order ID isn't found in the system; the "
        "Order ID matches more than one order and Return ID doesn't narrow it down; or Warehouse Comments "
        "are missing for a Damaged or Not Received row. The Image / Drive Link is recommended but not "
        "required — if it's left blank on a Damaged or Not Received row, the row still applies and the "
        "system adds an \"Image missing\" note to that return's comments. Inspection Status is no longer "
        "required on every row, and the same Order ID (or Order ID + Return ID) is allowed to repeat "
        "across multiple rows — each one is applied.")).alignment = Alignment(wrap_text=True)
    ins.row_dimensions[note_row].height = 55

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------
# Parsing + validation
# --------------------------------------------------------------------------

def _clean(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    s = str(v).strip()
    return "" if s.lower() == "nan" else s


def parse_uploaded_workbook(file) -> pd.DataFrame:
    """Reads the Bulk Inspection Upload sheet using row 1 (system field keys)
    as the header, skipping row 2 (the human-readable label row)."""
    df = pd.read_excel(file, sheet_name=0, header=0, skiprows=[1], dtype=object)
    df.columns = [str(c).strip() for c in df.columns]
    # keep only known columns, in our canonical order, tolerating extra/missing ones
    for col in SYSTEM_COLUMNS:
        if col not in df.columns:
            df[col] = None
    df = df[SYSTEM_COLUMNS]
    # drop fully blank rows
    df = df[~df.apply(lambda r: all(_clean(v) == "" for v in r), axis=1)]
    return df.reset_index(drop=True)


def _looks_like_mangled_number(s: str) -> bool:
    """Heuristic for the classic Excel/Sheets failure mode: a long ID pasted
    into a numeric-formatted cell becomes scientific notation (5.8506E+17) or
    gets rounded. Used only to give a more useful error message -- if the
    digits are already gone, we can't recover them, only explain why."""
    s = s.strip()
    if not s:
        return False
    if "e+" in s.lower() and any(ch.isdigit() for ch in s):
        return True
    # A long all-digit run ending in several zeros is also suspicious --
    # float64 only keeps ~15-17 significant digits, so a genuine 18-19 digit
    # ID that's been through a float round-trip often ends in trailing zeros.
    if s.isdigit() and len(s) >= 15 and s.endswith("000"):
        return True
    return False


def validate_upload(df: pd.DataFrame):
    """Returns (errors: list[dict], valid_rows: list[dict]).
    Each error dict: {row, order_id, field, message}.
    Each valid row dict carries the resolved ReturnOrder key plus cleaned
    values. A row whose status text wasn't recognized still comes back as a
    valid row (status_code=None, needs_review=True, status_raw preserved) --
    see the module docstring on STATUS_MAP for why that's applied rather than
    rejected."""
    errors = []
    valid_rows = []

    session = get_session()
    try:
        all_orders = session.query(
            ReturnOrder.key, ReturnOrder.order_id, ReturnOrder.return_id,
            ReturnOrder.trigger_code, ReturnOrder.order_status,
        ).all()
    finally:
        session.close()

    by_order_id = {}
    by_key = {}
    for key, order_id, return_id, trigger_code, order_status in all_orders:
        by_order_id.setdefault(order_id, []).append((key, return_id))
        by_key[key] = (trigger_code, order_status)

    for i, row in df.iterrows():
        excel_row = i + 4  # data starts at physical row 4 (after key/header/example rows)
        order_id = _clean(row["order_id"])
        return_id = _clean(row["return_id"])
        status_raw = _clean(row["inspection_status"])
        comments = _clean(row["warehouse_comments"])
        damage_reason = _clean(row["damage_reason"])
        link = _clean(row["image_drive_link"])
        insp_date_raw = row["inspection_date"]
        wh_user = _clean(row["warehouse_user_name"])

        row_errors = []

        if not order_id:
            row_errors.append(("order_id", "Order ID is required."))
        elif _looks_like_mangled_number(order_id):
            row_errors.append(("order_id",
                                f"'{order_id}' looks like it was converted to a number by Excel/Sheets and "
                                f"the original digits may be lost. Re-paste this Order ID as plain text from "
                                f"the source, or fix the source file (see the warning on the Instructions tab)."))

        # --- resolve which existing order this row is about ---
        resolved_key = None
        if order_id and not row_errors:
            matches = by_order_id.get(order_id)
            if not matches:
                row_errors.append(("order_id", "Order ID not found in the system."))
            elif len(matches) == 1:
                resolved_key = matches[0][0]
            else:
                if return_id:
                    narrowed = [k for k, rid in matches if rid == return_id]
                    if len(narrowed) == 1:
                        resolved_key = narrowed[0]
                    else:
                        row_errors.append(("order_id", "Multiple orders match this Order ID and "
                                                         "Return ID together — needs manual resolution."))
                else:
                    row_errors.append(("order_id", "Multiple orders share this Order ID — provide "
                                                     "Return ID to disambiguate."))
        # Note: an Order ID + Return ID pair repeating across rows in this file
        # is allowed on purpose -- the same order legitimately shows up more
        # than once when it has multiple SKUs or multiple return cases, and
        # each row just re-applies the same inspection result to that order.

        # --- resolve inspection status, honoring the "doesn't need inspection" exemption ---
        status_code = None
        needs_review = False
        trigger_code, order_status = by_key.get(resolved_key, (None, None))
        exempt_status = (CASES_NOT_REQUIRING_INSPECTION_BY_TRIGGER.get(trigger_code)
                          or CASES_NOT_REQUIRING_INSPECTION_BY_ORDER_STATUS.get(order_status))

        if status_raw:
            status_code = STATUS_MAP.get(status_raw.strip().lower())
            if status_code is None:
                # Unrecognized text (could be Thai phrasing we don't have yet,
                # a typo, or a status the dropdown doesn't list). Don't reject
                # the row for it -- apply it and keep the original text
                # visible so it can be reclassified by hand.
                needs_review = True
        elif exempt_status:
            # Blank is fine here -- this order's own data already shows it
            # never reached the warehouse for inspection.
            status_code = exempt_status
        elif resolved_key is not None:
            row_errors.append(("inspection_status",
                                "Inspection Status is required for this order (it doesn't already show as "
                                "cancelled or lost, so it needs an actual inspection result)."))
        # else: order_id itself didn't resolve -- the order_id error above already covers it.

        if status_code in STATUS_REQUIRES_EVIDENCE:
            if not comments:
                row_errors.append(("warehouse_comments",
                                    "Comments are required for Damaged / Not Received."))
            # Image/drive link is recommended but no longer mandatory -- a
            # missing link doesn't block the row (see note on
            # STATUS_REQUIRES_EVIDENCE above); it just gets flagged in the
            # comment below. A link that IS provided still has to look like
            # an actual http(s) link, though.
            if link and not is_probably_valid_link(link):
                row_errors.append(("image_drive_link",
                                    "This doesn't look like a valid http(s) link."))

        insp_date = None
        if insp_date_raw is not None and _clean(insp_date_raw):
            try:
                insp_date = pd.to_datetime(insp_date_raw).to_pydatetime()
            except Exception:
                row_errors.append(("inspection_date", f"'{insp_date_raw}' isn't a recognizable date."))

        if row_errors:
            for field, msg in row_errors:
                errors.append({"row": excel_row, "order_id": order_id or "(blank)",
                                "field": field, "message": msg})
        else:
            row_comments = comments
            if status_code in STATUS_REQUIRES_EVIDENCE and not link:
                img_note = "[Image missing at upload — follow up with warehouse for evidence]"
                row_comments = f"{img_note} {row_comments}".strip() if row_comments else img_note
            if needs_review:
                note = f"[Uploaded status not recognized — original text: \"{status_raw}\"]"
                row_comments = f"{note} {row_comments}".strip() if row_comments else note
            valid_rows.append({
                "row": excel_row, "key": resolved_key, "order_id": order_id, "return_id": return_id,
                "status_code": status_code, "status_raw": status_raw, "needs_review": needs_review,
                "comments": row_comments, "damage_reason": damage_reason,
                "link": link, "inspection_date": insp_date, "warehouse_user_name": wh_user,
            })

    return errors, valid_rows


def build_error_report_bytes(errors: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Errors"
    header_fill = PatternFill("solid", fgColor="B91C1C")
    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
    normal_font = Font(name=FONT_NAME, size=10)

    headers = ["Row", "Order ID", "Field", "Error"]
    ws.append(headers)
    for c in range(1, 5):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill

    for e in errors:
        ws.append([e["row"], e["order_id"], e["field"], e["message"]])
    for r in range(2, len(errors) + 2):
        for c in range(1, 5):
            ws.cell(row=r, column=c).font = normal_font

    widths = [8, 20, 22, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------
# Apply
# --------------------------------------------------------------------------

def apply_bulk_upload(valid_rows: list, fallback_user_name: str) -> dict:
    applied, skipped, flagged = 0, 0, 0
    session = get_session()
    try:
        for r in valid_rows:
            order = session.get(ReturnOrder, r["key"]) if r["key"] else None
            if order is None:
                skipped += 1
                continue
            who = r["warehouse_user_name"] or fallback_user_name
            when = r["inspection_date"] or dt.datetime.utcnow()

            # status_code is None only for a row whose status text wasn't
            # recognized (needs_review=True) -- leave inspection_status as
            # whatever it already was rather than nulling it out; the original
            # text is preserved in the comment (see validate_upload) for
            # someone to reclassify from the Returns page.
            if r["status_code"] is not None:
                order.inspection_status = r["status_code"]
            if r["comments"]:
                order.inspection_comment = r["comments"]
            if r["damage_reason"]:
                order.damage_reason = r["damage_reason"]
            if r["return_id"]:
                order.return_id = r["return_id"]
            order.inspected_by = who
            order.inspected_at = when
            session.commit()

            if r["link"]:
                save_link(order.key, r["link"], who, when=when)

            action = "Bulk inspection upload" if not r["needs_review"] else "Bulk inspection upload (needs review)"
            log_activity(order_key=order.key, action=action,
                          comment=f"Result: {r['status_code'] or r['status_raw']}"
                                  + (f" — {r['comments']}" if r["comments"] else ""))
            if r["status_code"] in STATUS_REQUIRES_EVIDENCE:
                push_notification("damaged", f"Order {order.order_id} needs Operations action "
                                              f"({r['status_code']})", order.key)
            if r["needs_review"]:
                flagged += 1
            applied += 1
    finally:
        session.close()
    return {"applied": applied, "skipped": skipped, "flagged": flagged}
